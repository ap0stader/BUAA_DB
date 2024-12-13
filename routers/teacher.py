from fastapi import FastAPI, Depends, HTTPException, status, APIRouter, File, UploadFile, Form
from fastapi.security import OAuth2PasswordBearer
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
import redis
import jwt
from datetime import datetime, timedelta, timezone
from typing import Optional
from utils.db import get_cursor
from utils.rds import rds
from io import BytesIO
import conf
import pandas as pd
from utils.errno import *
from utils.utils import *
from pathlib import Path
import os, uuid
import openpyxl
import tempfile
import random, string
from typing import List
from utils.resource import ResourceManager
from urllib.parse import quote

router = APIRouter(
    prefix="/Teacher",
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")


@router.get("/queryCurriculumScores")
async def query_curriculum_scores(curriculum_id: int):
    if not check_curriculum_id_exist(curriculum_id):
        return {
            'success': False,
            'errCode': 500101,
            'data': {}
        }
    conn, cursor = get_cursor('root')
    cursor.execute("""
        SELECT * FROM queryCurriculumScores qcs
        WHERE qcs.attendance_curriculum_id=%s;
        """, (curriculum_id,))
    scores = cursor.fetchall()
    return {
        'success': True,
        'errCode': 0,
        'data': {
            'scores': scores
        }
    }

@router.get("/downloadCurriculumScores")
async def download_curriculum_scores(curriculum_id: int):
    if not check_curriculum_id_exist(curriculum_id):
        return {
            'success': False,
            'errCode': 500201,
            'data': {}
        }
    conn, cursor = get_cursor('root')
    cursor.execute("""
        SELECT course_name, curriculum_id, curriculum_semester_name, curriculum_teacher_name AS teacher_name
        FROM curriculum_view
        WHERE curriculum_id=%s;
        """, (curriculum_id,))
    curriculum_info = cursor.fetchone()
    cursor.execute("""
        SELECT * FROM queryCurriculumScores qcs
        WHERE qcs.attendance_curriculum_id=%s;
        """, (curriculum_id,))
    scores = cursor.fetchall()
    xlsx_template = os.path.join("~", "static", "scores.xlsx")
    xlsx_template = os.path.expanduser(xlsx_template)
    wb = openpyxl.load_workbook(xlsx_template)
    ws = wb.active
    ws["B1"] = curriculum_info['course_name']
    ws["B2"] = curriculum_info['curriculum_id']
    ws["B3"] = curriculum_info['curriculum_semester_name']
    ws["B4"] = curriculum_info['teacher_name']
    for i, score in enumerate(scores):
        ws[f"A{i+6}"], ws[f"B{i+6}"], ws[f"C{i+6}"] = score['student_id'], score['student_name'], score['score']
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{curriculum_info['course_name']}_{curriculum_info['teacher_name']}_汇总表.xlsx"

    headers = {
        'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

class update_score_req(BaseModel):
    curriculum_id: int
    student_id: str
    score: float

@router.post("/updateScore")
async def update_score(req: update_score_req):
    curriculum_id, student_id, score = req.curriculum_id, req.student_id, req.score
    if not check_curriculum_id_exist(curriculum_id):
        return {
            'success': False,
            'errCode': 500301,
            'data': {}
        }
    if not check_is_student_in_curriculum(student_id, curriculum_id):
        return {
            'success': False,
            'errCode': 500302,
            'data': {}
        }
    conn, cursor = get_cursor('root')
    cursor.execute("""
        UPDATE attendance_table
        SET attendance_score=%s
        WHERE attendance_curriculum_id=%s AND attendance_student_id=%s;
        """, (score, curriculum_id, student_id))
    return {
        'success': True,
        'errCode': OK,
        'data': {}
    }

@router.post("/updateScoreBatch")
async def update_score_batch(
    curriculum_id: int = Form(...), 
    file: UploadFile = File(...)
):
    if not check_curriculum_id_exist(curriculum_id):
        return {
            'success': False,
            'errCode': 500401,
            'data': {}
        }
    conn, cursor = get_cursor('root')
    try:
        content = await file.read()
        df = pd.read_excel(BytesIO(content))
        for i in range(4, len(df)):
            student_id, score = df.iloc[i, 0], df.iloc[i, 2]
            # print(student_id, score)
            cursor.execute("""
                UPDATE attendance_table
                SET attendance_score=%s
                WHERE attendance_curriculum_id=%s AND attendance_student_id=%s;
                """, (score, curriculum_id, student_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return {
            'success': False,
            'errCode': 500402,
            'data': {}
        }
    return {
        'success': True,
        'errCode': OK,
        'data': {}
    }




